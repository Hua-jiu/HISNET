import os
import re
from cv2 import sort
import openpyxl as op

def extract_best_acc_from_file(file_path):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()
            if len(lines) < 2:
                print(f"File {file_path} does not have enough lines.")
                return None
            # 获取倒数第二行
            second_last_line = lines[-2]
            # 使用正则表达式查找 best acc 的值
            match = re.search(r"best acc: ([0-9\.]+)", second_last_line)
            if match:
                best_acc_value = float(match.group(1))
                return f"{best_acc_value:.2%}"
            else:
                print(f"Best acc not found in {file_path}.")
                return None
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return None

dir_path = '/mnt/storage-data2/anlong/MoleProject/New_data_Exp_20240410/Other_net_compare/To-Genus/Data_224'
weights_folder = f"{dir_path}/weights"
docs_folder = f"{dir_path}/docs"
best_acc_values = []

consequence_dict = {}
for file_name in os.listdir(docs_folder):
    if file_name.endswith(".txt"):
        file_path = os.path.join(docs_folder, file_name)
        best_acc = extract_best_acc_from_file(file_path)
        if best_acc is not None:
            file_name = file_name.split('.')[0]
            if file_name not in consequence_dict:
                consequence_dict[file_name] = {}
            consequence_dict[file_name]['best_acc'] = best_acc

# 可以对 best_acc_values 做进一步处理，比如打印或保存
work_file = op.Workbook()
work_sheet = work_file.active
model_name_list = []
for modelname in consequence_dict:
    model_name_list.append(modelname)
row, column = 1, 1
for model_name in sorted(model_name_list):
    work_sheet.cell(row, column).value  = model_name
    work_sheet.cell(row, column + 1).value = consequence_dict[model_name]['best_acc']
    work_sheet.cell(row, column + 2).value = 1
    work_sheet.cell(row, column + 3).value = 1
    row += 1
work_file.save(f"{docs_folder}/Z_best_acc.xlsx")
